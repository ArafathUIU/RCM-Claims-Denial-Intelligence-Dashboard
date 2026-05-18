"""RCM Claims Denial Intelligence - Excel Dashboard Builder.

Reads the generated claims CSV and produces a formatted
multi-sheet Excel workbook with charts and KPIs.
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, BarChart3D
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from copy import copy

# ============================================================
# CONFIGURATION
# ============================================================
DATA_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "claims_data.csv")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "RCM_Denial_Dashboard.xlsx")

# ============================================================
# STYLE CONSTANTS
# ============================================================
DARK_BLUE = "1B3A5C"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "BDD7EE"
WHITE = "FFFFFF"
RED_ACCENT = "C00000"
GREEN_ACCENT = "006100"
GRAY_BG = "F2F2F2"
ORANGE = "ED7D31"
TEAL = "2B9C8E"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=DARK_BLUE)
SUBTITLE_FONT = Font(name="Calibri", size=10, color="666666")
KPI_VALUE_FONT = Font(name="Calibri", size=24, bold=True, color=DARK_BLUE)
KPI_LABEL_FONT = Font(name="Calibri", size=10, color="888888")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
ALT_FILL = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type="solid")

CHART_COLORS = ["2E75B6", "ED7D31", "A5A5A5", "FFC000", "4472C4",
                "70AD47", "C00000", "5B9BD5", "264478", "636363"]


def apply_header_style(ws, row, cols, start_col=1):
    """Apply header styling to a row of cells."""
    for col in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def apply_body_style(ws, start_row, end_row, cols, start_col=1):
    """Apply alternating row colors and borders to data rows."""
    for r in range(start_row, end_row + 1):
        for c in range(start_col, start_col + cols):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if (r - start_row) % 2 == 1:
                cell.fill = ALT_FILL


def auto_width(ws, min_width=10, max_width=40):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)


def add_title(ws, title, row=1, col=1):
    """Add a sheet title."""
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = TITLE_FONT


def add_subtitle(ws, text, row=2, col=1):
    """Add a subtitle / description."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SUBTITLE_FONT


def add_kpi(ws, row, col, label, value, fmt=None):
    """Add a KPI card (label + value) at the given position."""
    v_cell = ws.cell(row=row, column=col, value=value)
    v_cell.font = KPI_VALUE_FONT
    v_cell.alignment = Alignment(horizontal="center")
    if fmt:
        v_cell.number_format = fmt

    l_cell = ws.cell(row=row + 1, column=col, value=label)
    l_cell.font = KPI_LABEL_FONT
    l_cell.alignment = Alignment(horizontal="center")


def build_summary_sheet(wb, df):
    """Sheet 1: Executive Summary with KPI cards and trend sparklines."""
    ws = wb.active
    ws.title = "Summary"

    add_title(ws, "RCM Claims Denial Intelligence", row=1, col=1)
    add_subtitle(ws, f"Data period: Jan 2024 - Dec 2025  |  Generated from {len(df):,} claims", row=2, col=1)

    total_claims = len(df)
    denied_df = df[df["claim_status"] != "Paid"]
    denied_count = len(denied_df)
    denial_rate = denied_count / total_claims * 100
    total_denied = denied_df["denied_amount"].sum()
    total_recovered = df[df["claim_status"] == "Recovered"]["recovered_amount"].sum()
    recovery_rate = total_recovered / total_denied * 100 if total_denied else 0
    appealed = denied_df[denied_df["appeal_flag"] == 1]
    appeal_win_rate = len(df[df["claim_status"] == "Recovered"]) / len(appealed) * 100 if len(appealed) else 0
    avg_aging = denied_df["aging_days"].mean()

    kpi_start_row = 5
    add_kpi(ws, kpi_start_row, 1, "Total Claims", total_claims, "#,##0")
    add_kpi(ws, kpi_start_row, 3, "Denial Rate", f"{denial_rate:.1f}%")
    add_kpi(ws, kpi_start_row, 5, "Total Denied", total_denied, "$#,##0")
    add_kpi(ws, kpi_start_row, 7, "Recovery Rate", f"{recovery_rate:.1f}%")
    add_kpi(ws, kpi_start_row + 3, 1, "Claims Denied", denied_count, "#,##0")
    add_kpi(ws, kpi_start_row + 3, 3, "Appeal Win Rate", f"{appeal_win_rate:.1f}%")
    add_kpi(ws, kpi_start_row + 3, 5, "Total Recovered", total_recovered, "$#,##0")
    add_kpi(ws, kpi_start_row + 3, 7, "Avg Aging Days", f"{avg_aging:.0f}")

    monthly = df.groupby(df["service_date"].str[:7]).agg(
        total=("claim_id", "count"),
        denied=("claim_status", lambda x: (x != "Paid").sum()),
        denied_amt=("denied_amount", "sum"),
        recovered_amt=("recovered_amount", "sum"),
    ).reset_index()
    monthly["denial_rate"] = monthly["denied"] / monthly["total"] * 100
    monthly.columns = ["Month", "Total Claims", "Denied Claims", "Denied Amount", "Recovered Amount", "Denial Rate %"]

    table_start = 11
    headers = list(monthly.columns)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=table_start, column=ci, value=h)
    apply_header_style(ws, table_start, len(headers))

    for ri, (_, row_data) in enumerate(monthly.iterrows()):
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=table_start + 1 + ri, column=ci, value=val)
        ws.cell(row=table_start + 1 + ri, column=4).number_format = "$#,##0"
        ws.cell(row=table_start + 1 + ri, column=5).number_format = "$#,##0"
        ws.cell(row=table_start + 1 + ri, column=6).number_format = "0.0"

    apply_body_style(ws, table_start + 1, table_start + len(monthly), len(headers))

    chart = LineChart()
    chart.title = "Monthly Denial Rate Trend"
    chart.y_axis.title = "Denial Rate %"
    chart.x_axis.title = "Month"
    chart.style = 10
    chart.height = 12
    chart.width = 22

    data_ref = Reference(ws, min_col=6, min_row=table_start, max_row=table_start + len(monthly))
    cats_ref = Reference(ws, min_col=1, min_row=table_start + 1, max_row=table_start + len(monthly))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.line.solidFill = MED_BLUE
    chart.series[0].graphicalProperties.line.width = 28000

    ws.add_chart(chart, f"A{table_start + len(monthly) + 3}")
    auto_width(ws)


def build_payer_sheet(wb, df):
    """Sheet 2: Payer Performance with bar chart."""
    ws = wb.create_sheet("Payer Performance")
    add_title(ws, "Payer Denial Performance", row=1, col=1)
    add_subtitle(ws, "Denial rate and denied amount by insurance payer", row=2, col=1)

    payer_stats = df.groupby("payer_name").agg(
        total=("claim_id", "count"),
        denied=("claim_status", lambda x: (x != "Paid").sum()),
        denied_amt=("denied_amount", "sum"),
        recovered_amt=("recovered_amount", "sum"),
    ).reset_index()
    payer_stats["denial_rate"] = payer_stats["denied"] / payer_stats["total"] * 100
    payer_stats["recovery_rate"] = (
        payer_stats["recovered_amt"] / payer_stats["denied_amt"].replace(0, None) * 100
    ).fillna(0)
    payer_stats = payer_stats.sort_values("denial_rate", ascending=False)
    payer_stats.columns = [
        "Payer", "Total Claims", "Denied Claims", "Denied Amount",
        "Recovered Amount", "Denial Rate %", "Recovery Rate %"
    ]

    headers = list(payer_stats.columns)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    apply_header_style(ws, 4, len(headers))

    for ri, (_, row_data) in enumerate(payer_stats.iterrows()):
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=5 + ri, column=ci, value=val)
    apply_body_style(ws, 5, 4 + len(payer_stats), len(headers))
    auto_width(ws, min_width=14)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Denial Rate % by Payer"
    chart.y_axis.title = "Denial Rate %"
    chart.style = 10
    chart.height = 14
    chart.width = 22

    data_ref = Reference(ws, min_col=6, min_row=4, max_row=4 + len(payer_stats))
    cats_ref = Reference(ws, min_col=1, min_row=5, max_row=4 + len(payer_stats))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = MED_BLUE

    ws.add_chart(chart, "I4")

    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Total Denied Amount by Payer"
    chart2.y_axis.title = "Denied $"
    chart2.style = 10
    chart2.height = 14
    chart2.width = 22

    data_ref2 = Reference(ws, min_col=4, min_row=4, max_row=4 + len(payer_stats))
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats_ref)
    chart2.series[0].graphicalProperties.solidFill = ORANGE

    ws.add_chart(chart2, "I20")


def build_denial_reasons_sheet(wb, df):
    """Sheet 3: Denial Reasons with Pareto chart."""
    ws = wb.create_sheet("Denial Reasons")
    add_title(ws, "Denial Reason Analysis", row=1, col=1)
    add_subtitle(ws, "Top denial codes by frequency and financial impact (Pareto)", row=2, col=1)

    denied = df[df["claim_status"] != "Paid"]
    reasons = denied.groupby("reason_code").agg(
        count=("claim_id", "count"),
        total_denied=("denied_amount", "sum"),
    ).reset_index()
    reasons = reasons.sort_values("total_denied", ascending=False)
    total_amt = reasons["total_denied"].sum()
    reasons["cumulative"] = reasons["total_denied"].cumsum()
    reasons["cum_pct"] = reasons["cumulative"] / total_amt * 100
    reasons["pct"] = reasons["total_denied"] / total_amt * 100

    headers = ["Reason Code", "Count", "Denied Amount", "% of Total", "Cumulative %"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    apply_header_style(ws, 4, len(headers))

    for ri, (_, row_data) in enumerate(reasons.iterrows()):
        ws.cell(row=5 + ri, column=1, value=row_data["reason_code"])
        ws.cell(row=5 + ri, column=2, value=row_data["count"])
        ws.cell(row=5 + ri, column=3, value=row_data["total_denied"])
        ws.cell(row=5 + ri, column=4, value=row_data["pct"])
        ws.cell(row=5 + ri, column=5, value=row_data["cum_pct"])
    apply_body_style(ws, 5, 4 + len(reasons), len(headers))
    auto_width(ws)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Denial Reasons Pareto Chart"
    chart.y_axis.title = "Denied Amount ($)"
    chart.style = 10
    chart.height = 14
    chart.width = 24

    data_ref = Reference(ws, min_col=3, min_row=4, max_row=4 + len(reasons))
    cats_ref = Reference(ws, min_col=1, min_row=5, max_row=4 + len(reasons))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = MED_BLUE

    line = LineChart()
    line_data = Reference(ws, min_col=5, min_row=4, max_row=4 + len(reasons))
    line.add_data(line_data, titles_from_data=True)
    line.y_axis.axId = 200
    line.series[0].graphicalProperties.line.solidFill = RED_ACCENT
    line.series[0].graphicalProperties.line.width = 25000
    line.series[0].marker.symbol = "circle"
    line.series[0].marker.size = 6

    chart.y_axis.crosses = "min"
    chart += line

    ws.add_chart(chart, "G4")


def build_recovery_sheet(wb, df):
    """Sheet 4: Revenue Recovery Analysis with waterfall chart."""
    ws = wb.create_sheet("Revenue Recovery")
    add_title(ws, "Revenue Recovery Analysis", row=1, col=1)
    add_subtitle(ws, "Appeals, recovered amounts, and recovery efficiency", row=2, col=1)

    denied = df[df["claim_status"] != "Paid"]
    total_billed = df["billed_amount"].sum()
    total_allowed = df["allowed_amount"].sum()
    total_denied_amt = denied["denied_amount"].sum()
    total_recovered_amt = df[df["claim_status"] == "Recovered"]["recovered_amount"].sum()
    total_written_off = total_denied_amt - total_recovered_amt

    stage_data = [
        ("Billed Charges", total_billed, ""),
        ("Allowed Amount", total_allowed, ""),
        ("Denied Amount", total_denied_amt, ""),
        ("Recovered", total_recovered_amt, ""),
        ("Written Off", total_written_off, ""),
    ]
    headers = ["Stage", "Amount", "Notes"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    apply_header_style(ws, 4, len(headers))

    for ri, (stage, amt, note) in enumerate(stage_data):
        ws.cell(row=5 + ri, column=1, value=stage)
        ws.cell(row=5 + ri, column=2, value=amt)
        ws.cell(row=5 + ri, column=2).number_format = "$#,##0"
        ws.cell(row=5 + ri, column=3, value=note)
    apply_body_style(ws, 5, 9, len(headers))
    auto_width(ws)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Revenue Flow: Billed → Denied → Recovered → Written Off"
    chart.y_axis.title = "Amount ($)"
    chart.style = 10
    chart.height = 14
    chart.width = 22

    data_ref = Reference(ws, min_col=2, min_row=4, max_row=9)
    cats_ref = Reference(ws, min_col=1, min_row=5, max_row=9)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    colors = [MED_BLUE, LIGHT_BLUE, RED_ACCENT, GREEN_ACCENT, ORANGE]
    for idx, color in enumerate(colors):
        chart.series[0].data_points.append(DataPoint(idx=idx, graphicalProperties_solidFill=color))

    ws.add_chart(chart, "E4")

    appeal_stats = denied.agg(
        total_denied=("claim_id", "count"),
        appeals=("appeal_flag", "sum"),
        wins=("claim_status", lambda x: (x == "Recovered").sum()),
        losses=("claim_status", lambda x: (x == "Written Off").sum()),
        denied_amt=("denied_amount", "sum"),
        recovered_amt=("recovered_amount", "sum"),
    )

    stats_data = [
        ("Total Denied Claims", appeal_stats["total_denied"], ""),
        ("Appeals Filed", appeal_stats["appeals"], f"{appeal_stats['appeals']/appeal_stats['total_denied']*100:.1f}%"),
        ("Appeals Won", appeal_stats["wins"], f"{appeal_stats['wins']/appeal_stats['appeals']*100:.1f}%" if appeal_stats["appeals"] else "0%"),
        ("Appeals Lost", appeal_stats["losses"], ""),
        ("Recovery Rate ($)", "", f"{appeal_stats['recovered_amt']/appeal_stats['denied_amt']*100:.1f}%" if appeal_stats["denied_amt"] else "0%"),
    ]
    stat_start = 12
    for ci, h in enumerate(["Metric", "Value", "Rate"], 1):
        ws.cell(row=stat_start, column=ci, value=h)
    apply_header_style(ws, stat_start, 3)
    for ri, (label, val, rate) in enumerate(stats_data):
        ws.cell(row=stat_start + 1 + ri, column=1, value=label)
        ws.cell(row=stat_start + 1 + ri, column=2, value=val)
        ws.cell(row=stat_start + 1 + ri, column=3, value=rate)
    apply_body_style(ws, stat_start + 1, stat_start + 5, 3)


def build_provider_sheet(wb, df):
    """Sheet 5: Provider & Department Drill-down."""
    ws = wb.create_sheet("Provider Drill-down")
    add_title(ws, "Provider & Department Performance", row=1, col=1)
    add_subtitle(ws, "Denial rates by provider and clinical department", row=2, col=1)

    provider_stats = df.groupby(["provider_name", "department"]).agg(
        total=("claim_id", "count"),
        denied=("claim_status", lambda x: (x != "Paid").sum()),
        denied_amt=("denied_amount", "sum"),
        recovered_amt=("recovered_amount", "sum"),
        avg_aging=("aging_days", "mean"),
    ).reset_index()
    provider_stats["denial_rate"] = provider_stats["denied"] / provider_stats["total"] * 100
    provider_stats["recovery_rate"] = (
        provider_stats["recovered_amt"] / provider_stats["denied_amt"].replace(0, None) * 100
    ).fillna(0)
    provider_stats = provider_stats.sort_values("denial_rate", ascending=False)
    provider_stats.columns = [
        "Provider", "Department", "Total Claims", "Denied Claims",
        "Denied Amount", "Recovered Amount", "Avg Aging Days",
        "Denial Rate %", "Recovery Rate %"
    ]

    headers = list(provider_stats.columns)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    apply_header_style(ws, 4, len(headers))

    from openpyxl.formatting.rule import CellIsRule

    for ri, (_, row_data) in enumerate(provider_stats.iterrows()):
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=5 + ri, column=ci, value=val)
        ws.cell(row=5 + ri, column=5).number_format = "$#,##0"
        ws.cell(row=5 + ri, column=6).number_format = "$#,##0"
        ws.cell(row=5 + ri, column=7).number_format = "0.0"
        ws.cell(row=5 + ri, column=8).number_format = "0.0%"
        ws.cell(row=5 + ri, column=9).number_format = "0.0%"

    last_row = 4 + len(provider_stats)
    apply_body_style(ws, 5, last_row, len(headers))

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_font = Font(color="9C0006")
    green_font = Font(color="006100")

    ws.conditional_formatting.add(
        f"H5:H{last_row}",
        CellIsRule(operator="greaterThan", formula=["15"], fill=red_fill, font=red_font)
    )
    ws.conditional_formatting.add(
        f"H5:H{last_row}",
        CellIsRule(operator="lessThanOrEqual", formula=["15"], fill=green_fill, font=green_font)
    )
    auto_width(ws, min_width=12)

    dept_stats = df.groupby("department").agg(
        total=("claim_id", "count"),
        denied=("claim_status", lambda x: (x != "Paid").sum()),
        denied_amt=("denied_amount", "sum"),
    ).reset_index()
    dept_stats["denial_rate"] = dept_stats["denied"] / dept_stats["total"] * 100
    dept_stats = dept_stats.sort_values("denial_rate", ascending=False)
    dept_stats.columns = ["Department", "Total Claims", "Denied", "Denied Amt", "Denial Rate %"]

    dept_start_col = len(headers) + 2
    dept_headers = list(dept_stats.columns)
    for ci, h in enumerate(dept_headers):
        ws.cell(row=4, column=dept_start_col + ci, value=h)
    apply_header_style(ws, 4, len(dept_headers), start_col=dept_start_col)

    for ri, (_, row_data) in enumerate(dept_stats.iterrows()):
        for ci, val in enumerate(row_data):
            ws.cell(row=5 + ri, column=dept_start_col + ci, value=val)
    apply_body_style(ws, 5, 4 + len(dept_stats), len(dept_headers), start_col=dept_start_col)

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Denial Rate % by Department"
    chart.style = 10
    chart.height = 12
    chart.width = 20

    data_ref = Reference(ws, min_col=dept_start_col + 4, min_row=4,
                         max_row=4 + len(dept_stats))
    cats_ref = Reference(ws, min_col=dept_start_col, min_row=5,
                         max_row=4 + len(dept_stats))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = TEAL

    chart_row = 4 + len(dept_stats) + 2
    ws.add_chart(chart, f"A{chart_row}")

    # Remove .gitkeep
    gk = os.path.join(os.path.dirname(__file__), ".gitkeep")
    if os.path.exists(gk):
        os.remove(gk)


def build_aging_sheet(wb, df):
    """Sheet 6: Aging & AR Analysis."""
    ws = wb.create_sheet("Aging & AR")
    add_title(ws, "Accounts Receivable Aging", row=1, col=1)
    add_subtitle(ws, "Denied claims aging distribution and outstanding AR", row=2, col=1)

    denied_df = df[df["claim_status"] != "Paid"].copy()
    denied_df["aging_bucket"] = pd.cut(
        denied_df["aging_days"],
        bins=[0, 30, 60, 90, float("inf")],
        labels=["0-30 days", "31-60 days", "61-90 days", "90+ days"],
    )

    aging = denied_df.groupby("aging_bucket", observed=False).agg(
        count=("claim_id", "count"),
        denied_amt=("denied_amount", "sum"),
        recovered_amt=("recovered_amount", "sum"),
        avg_aging=("aging_days", "mean"),
    ).reset_index()
    aging["outstanding"] = aging["denied_amt"] - aging["recovered_amt"]
    aging["pct"] = aging["denied_amt"] / aging["denied_amt"].sum() * 100
    aging.columns = ["Bucket", "Claims", "Denied Amt", "Recovered", "Avg Days", "Outstanding", "% Total"]
    aging_order = ["0-30 days", "31-60 days", "61-90 days", "90+ days"]
    aging["_sort"] = aging["Bucket"].apply(lambda x: aging_order.index(x) if x in aging_order else 99)
    aging = aging.sort_values("_sort").drop(columns=["_sort"])

    headers = list(aging.columns)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    apply_header_style(ws, 4, len(headers))

    for ri, (_, row_data) in enumerate(aging.iterrows()):
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=5 + ri, column=ci, value=val)
        ws.cell(row=5 + ri, column=3).number_format = "$#,##0"
        ws.cell(row=5 + ri, column=4).number_format = "$#,##0"
        ws.cell(row=5 + ri, column=5).number_format = "0.0"
        ws.cell(row=5 + ri, column=6).number_format = "$#,##0"
    apply_body_style(ws, 5, 8, len(headers))
    auto_width(ws, min_width=12)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Denied Claims by Aging Bucket"
    chart.y_axis.title = "Number of Claims"
    chart.style = 10
    chart.height = 14
    chart.width = 22

    data_ref = Reference(ws, min_col=2, min_row=4, max_row=8)
    cats_ref = Reference(ws, min_col=1, min_row=5, max_row=8)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    bucket_colors = [GREEN_ACCENT, "70AD47", ORANGE, RED_ACCENT]
    for idx, color in enumerate(bucket_colors):
        chart.series[0].data_points.append(DataPoint(idx=idx, graphicalProperties_solidFill=color))

    ws.add_chart(chart, "I4")

    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Outstanding AR by Aging Bucket"
    chart2.y_axis.title = "Amount ($)"
    chart2.style = 10
    chart2.height = 14
    chart2.width = 22

    data_ref2 = Reference(ws, min_col=6, min_row=4, max_row=8)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats_ref)
    for idx, color in enumerate(bucket_colors):
        chart2.series[0].data_points.append(DataPoint(idx=idx, graphicalProperties_solidFill=color))

    ws.add_chart(chart2, "I20")

    aging_by_payer = denied_df.groupby("payer_name").agg(
        avg_aging=("aging_days", "mean"),
        outstanding=("denied_amount", "sum"),
    ).reset_index()
    aging_by_payer = aging_by_payer.sort_values("avg_aging", ascending=False)
    aging_by_payer.columns = ["Payer", "Avg Aging Days", "Outstanding AR"]

    pay_start = 12
    for ci, h in enumerate(list(aging_by_payer.columns), 1):
        ws.cell(row=pay_start, column=ci, value=h)
    apply_header_style(ws, pay_start, len(aging_by_payer.columns))
    for ri, (_, row_data) in enumerate(aging_by_payer.iterrows()):
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=pay_start + 1 + ri, column=ci, value=val)
        ws.cell(row=pay_start + 1 + ri, column=2).number_format = "0.0"
        ws.cell(row=pay_start + 1 + ri, column=3).number_format = "$#,##0"
    apply_body_style(ws, pay_start + 1, pay_start + len(aging_by_payer), len(aging_by_payer.columns))
